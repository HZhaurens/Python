# Excel文件处理示例
import openpyxl

def create_excel(file_path):
    """创建Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入数据
    ws['A1'] = "姓名"
    ws['B1'] = "年龄"
    ws['A2'] = "张三"
    ws['B2'] = "25"
    
    wb.save(file_path)

def read_excel(file_path):
    """读取Excel文件内容"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    return data

# 示例用法
if __name__ == "__main__":
    # 创建Excel文件
    create_excel("example.xlsx")
    
    # 读取Excel文件
    content = read_excel("example.xlsx")
    for row in content:
        print(row)